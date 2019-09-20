import openpyxl
import mysql.connector 
# import pymysql
# import xlsxwriter
import pandas as pd
import numpy as np
import time
from config import db_connection_settings

#Gets GAIT data from MySQL database 
def GAIT_connect():

  mydb = mysql.connector.connect(**db_connection_settings, buffered=True)

  cursor = mydb.cursor()
  
  sql_query = ( 'SELECT '
       ' g.GrantID '
       ' , g.GrantTitle '
       ' , g.HQadmin '
       ' , g.USDAmount'
       ' , NULL as "Infra Budget"'
       ' , g.FundingStatus'
       ' , g.ProjectLength'
       ' , GROUP_CONCAT(DISTINCT ct.Country) as Country'
       ' , GROUP_CONCAT(DISTINCT rt.Region) as Region'
       ' , GROUP_CONCAT(DISTINCT dt.Donor) as Donor'
       ' , GROUP_CONCAT(DISTINCT dpt.Department) as DonorDept'
       ' , GROUP_CONCAT(DISTINCT mt.Methodology) as Methodology'
       ' , GROUP_CONCAT(DISTINCT s.Sector) as Sector'
       ' , GROUP_CONCAT(DISTINCT ss.SubSector) as "Area of Focus"'
       ' , GROUP_CONCAT(DISTINCT cct.CostCenter)  as "Fund Code"'
       ' , g.FundingProbability as FundingProbability'
       ' , CASE WHEN g.ComplexProgram = 1 THEN "Complex" ELSE "Not Complex" END AS "Complex Program"'
       ' , g.StartDate as StartDate'
       ' , g.EndDate as EndDate' 
       ' FROM granttbl g LEFT JOIN grantcountrytbl gct'
       ' ON g.GrantID=gct.GrantID LEFT JOIN countrytbl ct'
       ' ON gct.CountryID=ct.CountryID LEFT JOIN n_grantmethodologytbl gmt'
       ' ON g.GrantID=gmt.GrantID LEFT JOIN n_methodologytbl mt'
       ' ON gmt.MethodologyID=mt.MethodologyID LEFT JOIN costcentertbl cct'
       ' ON g.GrantID=cct.GrantID LEFT JOIN regiontbl rt'
       ' ON ct.RegionID=rt.RegionID LEFT JOIN donortbl dt'
       ' ON g.DonorID=dt.DonorID LEFT JOIN donordepttbl dpt'
       ' ON g.DepartmentID=dpt.DepartmentID LEFT JOIN n_grantsectortbl gs'
       ' ON g.GrantID = gs.GrantID LEFT JOIN n_subsectortbl ss'
       ' ON gs.SubSectorID = ss.SubSectorID LEFT JOIN n_sectortbl s'
       ' ON s.SectorID = gs.SectorID'
       ' WHERE (FundingStatus = "Closed" OR FundingStatus = "Completed")'
       ' AND (year(g.EndDate) >= 2016)'
       ' GROUP BY g.GrantID'
       ' ORDER BY g.GrantTitle')

  cursor.execute(sql_query)
  df = pd.read_sql_query(sql_query, mydb)
  df['Fund Code'].replace(to_replace=[None], value = np.nan, inplace=True)
  df.astype({'Fund Code':'str'}).dtypes
  mydb.close()
  return df


#Searches through financial sheet for rows with a description containing any of the infra 'key words'  
def search(excel_file, sheet_name):

  spreadsheet = openpyxl.load_workbook(excel_file)

  foundRows = []

  #Source Sheet: Where I pasted the financial sheet
  sheet = spreadsheet[sheet_name]

  #Destination Sheet: Outputs all rows that have infra key words in their description
  paste_sheet = spreadsheet['Sheet2']

  #Copy and paste the column names (Fund Number, Description, Local Currency Amount) from the financial sheet 
  paste_sheet.cell(row = 1, column = 1).value = sheet.cell(row = 1, column = 4).value
  paste_sheet.cell(row = 1, column = 2).value = sheet.cell(row = 1, column = 14).value
  paste_sheet.cell(row = 1, column = 3).value = sheet.cell(row = 1, column = 15).value

  #These are the key words that are being search in the 'Description' from the Financial Sheet
  key_words = [
    'latrine', 'borehole', 'waterpoint', 'water point', 'repair', 'rehabilitation',
    'construct', 'build', 'rehab', 'const', 'hand pump', 'rehab.', 'renov', 'septic tank', 
    'supply', 'install', 'renovation', 'pump station', 'sport field', 'forage', 'drilling',
    'water tank', 'watertank', 'water network', 'health center', 'school', 'pipe', 'hospital',
    'clinic', 'solar pump', 'water system'
  ]

  for row in range(1, sheet.max_row):
    if sheet.cell(row, 14).value is not None:
      for word in key_words:
        if ((word in sheet.cell(row, 14).value) or (word.capitalize() in sheet.cell(row, 14).value) or (word.upper() in sheet.cell(row, 14).value)) and (sheet.cell(row,15).value > 1000):
          foundRows.append(row)
          break;

  #Paste actual rows that have infra keywords in their description
  i = 2 
  for j in foundRows:
    paste_sheet.cell(row = i, column = 1).value = sheet.cell(row = j, column = 4).value
    paste_sheet.cell(row = i, column = 2).value = sheet.cell(row = j, column = 14).value
    paste_sheet.cell(row = i, column = 3).value = sheet.cell(row = j, column = 15).value
    i = i + 1 

  spreadsheet.save(excel_file)

#Helper function which clears excel sheets
def clear_sheet(excel_file):
  spreadsheet = openpyxl.load_workbook(excel_file)
  sheet_two = spreadsheet['Sheet2']
  sheet_three = spreadsheet['Sheet3']

  for row in sheet_two['A1:Z10000']:
    for cell in row:
      cell.value = None

  for row in sheet_three['A1:Z10000']:
    for cell in row:
      cell.value = None

  spreadsheet.save(excel_file) 

#Groups duplicate fund codes and sums their local currency amount. 
def group(excel_file):
  df = pd.read_excel(excel_file, sheet_name = 'Sheet2')
  grouped = df.groupby('Fund Number', as_index = False).agg({"Local Currency Amount": "sum"})
  return grouped 

#Helper function: writes the GAIT data to one sheet, and the grouped infra fundcodes in another sheet
def get_data(excel_file):
  
  df1 = GAIT_connect()
  df2 = group(excel_file)

  writer = pd.ExcelWriter(excel_file, engine = 'openpyxl')
  df1.to_excel(writer, sheet_name = 'Sheet3', index = False)
  df2.to_excel(writer, sheet_name = 'Sheet2', index = False)
  writer.save()

#Cross-references the GAIT data with the grouped data. If a fund code is in both sheets, copy the infrastructure budget (the summed 'Local Currency Amount')
def compare(excel_file):
  found_rows = []
  infra_budget = []

  spreadsheet = openpyxl.load_workbook(excel_file)
  sheet_two = spreadsheet['Sheet2']
  sheet_three = spreadsheet['Sheet3']

  group = pd.read_excel(excel_file, sheet_name = 'Sheet2')
  gait = pd.read_excel(excel_file, sheet_name = 'Sheet3')

  df1 = pd.DataFrame(group)

  df2 = pd.DataFrame(gait)

  gait_list = df2['Fund Code'].tolist()
  group_list = df1['Fund Number'].tolist()
  gait_list = list(map(str, gait_list))
  group_list = list(map(str, group_list))


  for i in range(2, sheet_three.max_row):
      if sheet_three.cell(i, 15).value is not None:
        for fund_code in group_list:
            if fund_code in sheet_three.cell(i, 15).value:
                found_rows.append(i)
                infra_budget.append(group_list.index(fund_code)+2)



  for i, j in zip(found_rows, infra_budget):
    if sheet_three.cell(i, 5).value is not None:
      sheet_three.cell(i, 5).value += sheet_two.cell(j, 2).value
    else:
      sheet_three.cell(i, 5).value = sheet_two.cell(j, 2).value


  spreadsheet.save(excel_file)

#Copies the Financial Sheet to a sheet in the consolidated workbook
def copyData(financial_sheet, excel_file, sheet_name):
  wb1 = openpyxl.load_workbook(financial_sheet, data_only = True)
  ws1 = wb1['Detail']

  wb2 = openpyxl.load_workbook(excel_file)

  if sheet_name not in wb2.sheetnames:
    wb2.create_sheet(sheet_name)

  ws2 = wb2[sheet_name]

  for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

  ws2.delete_cols(1, 3)
  ws2.delete_rows(1, 7)
  wb2.save(excel_file) 

      
def main():
  print("~~~~PROGRAM EXECUTING~~~~")
  excel_file = 'consolidated_infra_file.xlsx'
  financial_sheet = '8400 Infrastructure costs.xlsx'
  sheet_name = 'financial_data'
  copyData(financial_sheet, excel_file, sheet_name)
  clear_sheet(excel_file)
  search(excel_file, sheet_name)
  get_data(excel_file)
  compare(excel_file)
  print("~~~~ALL DONE~~~~")

  
if __name__== "__main__":
  main()


